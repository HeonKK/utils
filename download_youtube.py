
from pytube import YouTube
import requests
import subprocess
import os
from openpyxl import load_workbook

file_path = './data_collection.xlsx'
sheet_name = 'Korean'

load_wb = load_workbook(file_path, data_only=True)

load_ws = load_wb[sheet_name]

f = open('video_list.txt', 'a')


row = 3
col = 2
print(load_ws.cell(row, col).value)
while (load_ws.cell(row, col).value):
    
    url = str(load_ws.cell(row, col).value)
            
    index = load_ws.cell(row, 1).value
    
    workdir = './Downloads/video/'
    
    #url = 'https://www.youtube.com/watch?v=bps3m4eFTuE'
    
        
    yt = YouTube(url)
    file_name = load_ws.cell(row, 3).value
    print('Downloading %s' % (file_name))
    res = 720
    stream = yt.streams.filter(res='720p', progressive=True).first()    
    if stream is None:
        stream = yt.streams.filter(res='480p', progressive=True).first()
        print("480p\n")
        res = 480
        if stream is None:
            stream = yt.streams.filter(res='360p', progressive=True).first()
            print("360p\n")
            res = 360
    
            if stream is None:
                print('\nCan\'t find stream!\n')
                row+=1
                continue

    
    
    f.write('{:d} {:s} {:d} {:.1f}\n'.format(index, file_name, res, stream.fps))
    
    stream.download(workdir+file_name)
    print('Completed')
    
    row+=1
    if row == 4:
        break
        
f.close()
